"""FastAPI router for exporting generated content as Excel or Markdown.

Endpoints:
  GET /api/v1/content/export/{package_id}?format=xlsx|md
"""
from __future__ import annotations

import io
import logging

from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session

from app.database import get_db
from app.models.sqlalchemy_models import ContentPackage
from fastapi import HTTPException

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/api/v1/content", tags=["content"])


@router.get("/export-by-task/{task_id}")
async def export_content_by_task(
    task_id: str,
    format: str = Query("md", pattern="^(md|xlsx)$"),
    db: Session = Depends(get_db),
):
    """通过任务ID导出生成的内容"""
    pkg = db.query(ContentPackage).filter(ContentPackage.task_id == task_id).first()
    if not pkg:
        raise HTTPException(status_code=404, detail="内容包不存在")
    if not pkg.result:
        raise HTTPException(status_code=400, detail="内容尚未生成，无法导出")

    if format == "xlsx":
        return _export_xlsx(pkg)
    else:
        return _export_markdown(pkg)


@router.get("/export/{package_id}")
async def export_content_package(
    package_id: str,
    format: str = Query("md", pattern="^(md|xlsx)$"),
    db: Session = Depends(get_db),
):
    """一键导出生成的内容（支持 Markdown / Excel 格式）"""
    pkg = db.query(ContentPackage).filter(ContentPackage.id == package_id).first()
    if not pkg:
        raise HTTPException(status_code=404, detail="内容包不存在")
    if not pkg.result:
        raise HTTPException(status_code=400, detail="内容尚未生成，无法导出")

    if format == "xlsx":
        return _export_xlsx(pkg)
    else:
        return _export_markdown(pkg)


def _export_markdown(pkg: ContentPackage) -> StreamingResponse:
    """导出为 Markdown 文件"""
    result = pkg.result or {}
    angles = result.get("angles", [])
    posts = result.get("posts", [])

    lines = [
        f"# 内容导出报告",
        f"",
        f"**行业：** {pkg.industry}  **风格：** {pkg.style}",
        f"**生成时间：** {pkg.created_at.strftime('%Y-%m-%d %H:%M') if pkg.created_at else '未知'}",
        f"",
        "---",
        f"",
        "## 爆款切入点",
        "",
    ]

    for i, angle in enumerate(angles, 1):
        lines.extend([
            f"### 角度 {i}：{angle.get('title', '')}",
            f"",
            f"> {angle.get('hook', '')}",
            f"",
            f"{angle.get('angle_description', '')}",
            f"",
        ])

    lines.extend(["---", "", "## 各平台内容", ""])

    for post in posts:
        platform = post.get("platform", "未知平台")
        lines.extend([
            f"### {platform}",
            f"",
            f"**标题：** {post.get('title', '')}",
            f"",
            f"{post.get('body', '')}",
            f"",
            f"**话题标签：** {' '.join(post.get('tags', []))}",
            f"",
            f"**配图建议：** {post.get('visual_suggestion', '')}",
            f"",
            "---",
            "",
        ])

    content = "\n".join(lines)
    filename = f"内容导出_{pkg.industry}_{pkg.created_at.strftime('%Y%m%d') if pkg.created_at else 'unknown'}.md"

    return StreamingResponse(
        iter([content.encode("utf-8")]),
        media_type="text/markdown; charset=utf-8",
        headers={
            "Content-Disposition": f'attachment; filename="{filename}"',
        },
    )


def _export_xlsx(pkg: ContentPackage) -> StreamingResponse:
    """导出为 Excel (.xlsx) 文件"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    result = pkg.result or {}
    angles = result.get("angles", [])
    posts = result.get("posts", [])

    wb = Workbook()

    # --- Sheet 1：内容概览 ---
    ws1 = wb.active
    ws1.title = "内容概览"

    header_font = Font(name="微软雅黑", bold=True, size=12, color="FFFFFF")
    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    body_font = Font(name="微软雅黑", size=10)
    body_align = Alignment(wrap_text=True, vertical="top")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # 头部信息
    ws1.append(["矩阵内容工厂 - 内容导出报告"])
    ws1.merge_cells("A1:E1")
    ws1["A1"].font = Font(name="微软雅黑", bold=True, size=16, color="2563EB")

    ws1.append([f"行业：{pkg.industry}  |  风格：{pkg.style}  |  平台：{'、'.join(pkg.platforms) if isinstance(pkg.platforms, list) else pkg.platforms}"])
    ws1.merge_cells("A2:E2")

    ws1.append([])  # 空行

    # 角度表头
    angle_headers = ["序号", "爆款标题", "开头钩子句", "角度解析"]
    ws1.append(angle_headers)
    for col in range(1, 5):
        cell = ws1.cell(row=4, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    for i, angle in enumerate(angles, 1):
        ws1.append([
            i,
            angle.get("title", ""),
            angle.get("hook", ""),
            angle.get("angle_description", ""),
        ])
        for col in range(1, 5):
            cell = ws1.cell(row=4 + i, column=col)
            cell.font = body_font
            cell.alignment = body_align
            cell.border = thin_border

    ws1.column_dimensions["A"].width = 8
    ws1.column_dimensions["B"].width = 40
    ws1.column_dimensions["C"].width = 50
    ws1.column_dimensions["D"].width = 50

    # --- Sheet 2：内容明细 ---
    ws2 = wb.create_sheet("内容明细")

    detail_headers = ["平台", "标题", "正文", "话题标签", "配图建议"]
    ws2.append(detail_headers)
    for col in range(1, 6):
        cell = ws2.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    for i, post in enumerate(posts, 2):
        tags_str = " ".join(post.get("tags", []))
        ws2.append([
            post.get("platform", ""),
            post.get("title", ""),
            post.get("body", ""),
            tags_str,
            post.get("visual_suggestion", ""),
        ])
        for col in range(1, 6):
            cell = ws2.cell(row=i, column=col)
            cell.font = body_font
            cell.alignment = body_align
            cell.border = thin_border

    ws2.column_dimensions["A"].width = 15
    ws2.column_dimensions["B"].width = 40
    ws2.column_dimensions["C"].width = 60
    ws2.column_dimensions["D"].width = 40
    ws2.column_dimensions["E"].width = 40
    # Set body column to wrap text
    for row in ws2.iter_rows(min_row=2, max_col=3):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    # 输出
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"内容导出_{pkg.industry}_{pkg.created_at.strftime('%Y%m%d') if pkg.created_at else 'unknown'}.xlsx"

    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f'attachment; filename="{filename}"',
        },
    )
